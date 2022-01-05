# -*- coding: utf-8 -*-
import sys
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

import smtplib # 메일을 보내기 위한 라이브러리 모듈
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

import chromedriver_autoinstaller
import time
import subprocess
import shutil
import openpyxl
import datetime
import yaml
import os
import argparse

import pyautogui

import requests
from google_drive_downloader import GoogleDriveDownloader as gdd


parser = argparse.ArgumentParser(description='Automate smart store.')
parser.add_argument('--conf', default='./conf/conf.yaml', help='configuration file')
args = parser.parse_args()


def initWeb (chromePathList):
    try:
        shutil.rmtree(r"c:\chrometemp")  #쿠키 / 캐쉬파일 삭제
    except FileNotFoundError:
        pass

    for path in chromePathList:
        command = fr'{path} --remote-debugging-port=9222 --user-data-dir="C:\chrometemp"'
        try:
            subprocess.Popen(command) # 디버거 크롬 구동
            break
        except:
            print("")

    # option = Options()
    option = webdriver.ChromeOptions()
    option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

    chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
    try:
        # driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options=option)
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=option)
    except:
        chromedriver_autoinstaller.install(True)
        # driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe', options=option)
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=option)
    driver.implicitly_wait(10)

    return driver


def getOderInfoFromStore(driver, smartStoreConf):

    url = 'https://nid.naver.com/nidlogin.login?url=https%3A%2F%2Fsell.smartstore.naver.com%2F%23%2FnaverLoginCallback%3Furl%3Dhttps%253A%252F%252Fsell.smartstore.naver.com%252F%2523'
    driver.get(url)

    id = smartStoreConf['smartStoreId']
    pswd = smartStoreConf['smartStorePass']

    driver.find_element(By.ID, 'id').send_keys(id)
    driver.find_element(By.ID, 'pw').send_keys(pswd)
    driver.find_element(By.CLASS_NAME, 'btn_text').click()
    #print(driver.page_source)

    # 발주(주문)확인/발송관리 페이지
    driver.get('https://sell.smartstore.naver.com/#/naverpay/sale/delivery')
    # driver.implicitly_wait(5)
    # driver.find_element(By.XPATH, '//*[@id="seller-lnb"]/div/div[1]/ul/li[2]/a').click()
    # driver.find_element(By.XPATH, '//*[@id="seller-lnb"]/div/div[1]/ul/li[2]/ul/li[4]/a').click()

    # 주문목록 프레임으로 변경
    driver.switch_to.frame("__delegate")

    table  = driver.find_element(By.XPATH, '//*[@id="__app_root__"]/div/div[2]/div[4]/div[4]/div[1]/div[2]/div[1]/div[1]/div[2]/div/div[1]/table/tbody')
    table2 = driver.find_element(By.XPATH, '//*[@id="__app_root__"]/div/div[2]/div[4]/div[4]/div[1]/div[2]/div[1]/div[2]/div[2]/div/div[1]/table/tbody')

    orderInfoList = []

    for i, tr in enumerate(table.find_elements(By.TAG_NAME, "tr")):

        orderInfo = {}

        # 우편번호
        tr2 = table2.find_elements(By.TAG_NAME, "tr")
        td2 = tr2[i].find_elements(By.TAG_NAME, "td")[7]
        e = td2.find_element(By.CLASS_NAME, 'tui-grid-cell-content')
        orderInfo['zip_code'] = e.text

        # 상품주문번호
        td = tr.find_elements(By.TAG_NAME, "td")[1]
        e = td.find_element(By.CLASS_NAME, 'tui-grid-cell-content')
        orderInfo['order_num'] = e.text
        # driver.implicitly_wait(5)
        e.click()

        # 팝업창으로 변경
        driver.switch_to.window(driver.window_handles[1])

        # 상품명
        e = driver.find_element(By.XPATH, '//*[@id="__app_root__"]/div/div/div[2]/div[1]/div/div/div[2]/table/tbody/tr[1]/td')
        orderInfo['product_name'] = e.text

        # 구매자명
        e = driver.find_element(By.XPATH, '//*[@id="__app_root__"]/div/div/div[2]/div[1]/div/div/div[2]/table/tbody/tr[3]/td[1]')
        orderInfo['buy_name'] = e.text

        # 옵션
        e = driver.find_element(By.XPATH, '//*[@id="__app_root__"]/div/div/div[2]/div[1]/div/div/div[2]/table/tbody/tr[5]/td[1]')
        orderInfo['option'] = e.text

        # 주문수량
        e = driver.find_element(By.XPATH, '//*[@id="__app_root__"]/div/div/div[2]/div[1]/div/div/div[2]/table/tbody/tr[5]/td[2]')
        orderInfo['count'] = e.text

        # 수취인명
        e = driver.find_element(By.XPATH, '//*[@id="__app_root__"]/div/div/div[2]/div[1]/div/div/div[4]/table/tbody/tr[1]/td')
        orderInfo['to_name'] = e.text

        # 연락처1
        e = driver.find_element(By.XPATH, '//*[@id="__app_root__"]/div/div/div[2]/div[1]/div/div/div[4]/table/tbody/tr[2]/td[1]')
        orderInfo['to_call_num_1'] = e.text

        # 연락처2
        e = driver.find_element(By.XPATH, '//*[@id="__app_root__"]/div/div/div[2]/div[1]/div/div/div[4]/table/tbody/tr[2]/td[2]')
        orderInfo['to_call_num_2'] = e.text

        # 배송지
        e = driver.find_element(By.XPATH, '//*[@id="__app_root__"]/div/div/div[2]/div[1]/div/div/div[4]/table/tbody/tr[3]/td')
        orderInfo['to_add'] = e.text

        # 배송메모
        e = driver.find_element(By.XPATH, '//*[@id="__app_root__"]/div/div/div[2]/div[1]/div/div/div[4]/table/tbody/tr[4]/td')
        orderInfo['to_message'] = e.text

        orderInfoList.append(orderInfo)

        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        driver.switch_to.frame("__delegate")

    driver.close()

    return orderInfoList


def getOrderInfoList_jenia (orderInfoList):

    product_info_wb = openpyxl.load_workbook('등록제품 정보.xlsx')
    jn_sheet = product_info_wb['제니아']

    orderInfoList_jenia = []
    for orderInfo in orderInfoList:
        product_name, price = get_product_name_and_price (jn_sheet, orderInfo['product_name'], orderInfo['option'])
        if len(product_name) == 0:
            continue
        orderInfo_jenia = orderInfo
        orderInfo_jenia['jenia_product_name'] = product_name
        orderInfo_jenia['jenia_price'] = int(price)
        orderInfoList_jenia.append(orderInfo_jenia)

    orderInfoList_jenia.sort(key=lambda orderInfo: [orderInfo['jenia_product_name'], orderInfo['order_num']])
    return orderInfoList_jenia

def get_product_name_and_price (jn_sheet, order_product_name, order_option):

    # print(order_product_name)
    # print(order_option)

    # 주요 컬럼 위치 탐색
    # 제품명, 모델명, 벤더 공급가액, 네이버 제품명,	네이버 옵션
    title = ["제품명", "모델명", "벤더 공급가액", "네이버 제품명", "네이버 옵션"]

    titleDict = {}
    for row in range(1,20):
        for col in range(1,50):
            cell_value = jn_sheet.cell(row, col).value
            if cell_value in title:
                titleDict[cell_value] = [row, col]

            if len(titleDict) == len(title):
                break

    if len(titleDict) == 0:
        return "", 0

    row = titleDict[title[0]][0] + 1
    product_name = ''
    price = 0
    while True:
        jn_product_name = jn_sheet.cell(row, titleDict['제품명'][1]).value
        jn_model = jn_sheet.cell(row, titleDict['모델명'][1]).value
        jn_price = jn_sheet.cell(row, titleDict['벤더 공급가액'][1]).value

        jn_naver_product_name = jn_sheet.cell(row, titleDict['네이버 제품명'][1]).value
        jn_naver_option = jn_sheet.cell(row, titleDict['네이버 옵션'][1]).value

        if jn_product_name is None:
            break
        # if jn_naver_product_name is not None:
        #     print(jn_naver_product_name)
        #     print(jn_naver_option)
        if jn_naver_product_name == order_product_name and jn_naver_option in order_option:
            if jn_model is None:
                product_name = jn_product_name
            else:
                product_name = jn_model
            price = jn_price
            break
        row += 1
        if row > 500:
            break
        # print (row)

    return product_name, price

def saveOderXlsx_jenia (orderInfoList, formExcelFilename, purchaseOrderExcelFilename):
    wb = openpyxl.load_workbook(formExcelFilename)
    sheet = wb['발주서양식']

    d = datetime.datetime.now()
    sheet['C4'].value = d.strftime("%Y-%m-%d")

    order_dict = {}

    total_price = 0
    xlx_row = 8
    for orderInfo in orderInfoList:
        product_name = orderInfo['jenia_product_name']
        price = orderInfo['jenia_price']

        total_price += price * int(orderInfo['count'])
        sheet['A'+str(xlx_row)].value = orderInfo['jenia_product_name']
        sheet['C'+str(xlx_row)].value = int(orderInfo['count'])
        sheet['f'+str(xlx_row)].value = orderInfo['to_name']
        sheet['g'+str(xlx_row)].value = orderInfo['to_call_num_1']
        sheet['h'+str(xlx_row)].value = orderInfo['to_call_num_2']
        sheet['i'+str(xlx_row)].value = orderInfo['zip_code']
        sheet['j'+str(xlx_row)].value = orderInfo['to_add'].replace("\n", " ")
        sheet['k'+str(xlx_row)].value = orderInfo['to_message']

        if product_name in order_dict:
            order_dict[product_name][0] = order_dict[product_name][0] + int(orderInfo['count'])
        else:
            order_dict[product_name] = [int(orderInfo['count']), price]
        xlx_row += 1

    wb.save(purchaseOrderExcelFilename)

    return order_dict


def genMailMsg_jenia (order_summary_dict):
    d = datetime.datetime.now()

    total_price = 0
    mail_product_info = ""
    for order in order_summary_dict:
        cur_price = order_summary_dict[order][0] * order_summary_dict[order][1]
        total_price += cur_price
        msg = f"{order} {order_summary_dict[order][0]}개 ({format(cur_price, ',')}원)"
        mail_product_info += msg + '\n'


    mail_msg = f"""
안녕하세요?
더티키 입니다.
{d.strftime("%m월 %d일")} 발주서 입니다.

{mail_product_info}
더티키박남정으로 {format(total_price, ",")}원 입금 했습니다.

즐거운 하루 보내세요.
    """

    return mail_msg


def sendMail (id, pawd, toEmail, mail_msg, purchaseOrderFilename):

    d = datetime.datetime.now()

    sendEmail = id
    recvEmail = toEmail
    password = pawd
    smtpName = "smtp.gmail.com"
    smtpPort = 587

    #여러 MIME을 넣기위한 MIMEMultipart 객체 생성
    msg = MIMEMultipart()
    msg['Subject'] = f'{d.strftime("%m월 %d일")} 더티키 발주서 입니다.'
    msg['From'] = sendEmail
    msg['To'] = recvEmail

    #본문 추가
    contentPart = MIMEText(mail_msg) #MIMEText(text , _charset = "utf8")
    msg.attach(contentPart)

    #파일 추가
    etcFileName = purchaseOrderFilename
    with open(etcFileName, 'rb') as etcFD :
        etcPart = MIMEApplication( etcFD.read() )
        #첨부파일의 정보를 헤더로 추가
        etcPart.add_header('Content-Disposition','attachment', filename=etcFileName)
        msg.attach(etcPart)

    s=smtplib.SMTP( smtpName , smtpPort )
    s.starttls()
    s.login( sendEmail , password )
    s.sendmail( sendEmail, recvEmail, msg.as_string() )
    s.close()


def main():
    with open(args.conf, 'r', encoding="utf-8") as stream:
        try:
            conf = yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(exc)

    os.makedirs(conf['jenia']['purchaseOrderPath'], exist_ok=True)

    # 구글드라이브에서 등록제품 정보 다운로드
    gdd.download_file_from_google_drive(file_id=conf['googleDriveFileId'],
                                    dest_path='./등록제품 정보.xlsx',
                                    unzip=True)

    # 스마트스토어에서 주문정보 가져오기
    driver = initWeb(conf['chromePath'])
    maxCount = conf['maxTry']
    for countTry in range(maxCount):
        if countTry >= maxCount:
            break
        print(f"try get oder info from store: {countTry + 1} th")
        try:
            orderInfoList = getOderInfoFromStore(driver, conf['smartStore'])
            break
        except:
            print("예외 발생")

    if len(orderInfoList) == 0:
        print('주문목록이 없습니다.')
        sys.exit()

    # test data 저장
    # with open("orderinfo.yaml", 'w', encoding="utf-8") as stream:
    #     yaml.dump(orderInfoList, stream, allow_unicode=True)
    # test data 로드
    # with open("orderinfo.yaml", 'r', encoding="utf-8") as stream:
    #     orderInfoList = yaml.safe_load(stream)

    # 제니아 주문만 가져오기
    orderInfoList_jenia = getOrderInfoList_jenia(orderInfoList)

    if len(orderInfoList_jenia) > 0:
        d = datetime.datetime.now()
        dd = d.strftime("%Y%m%d")

        purchaseOrderFilename = os.path.join(conf['jenia']['purchaseOrderPath'], f"{dd}_더티키_발주서.xlsx")

        order_summary_dict = saveOderXlsx_jenia (
                                    orderInfoList_jenia,
                                    conf['jenia']['purchaseOrderFormFile'],
                                    purchaseOrderFilename)
        mail_msg = genMailMsg_jenia(order_summary_dict)

        print(mail_msg)
        if conf['mail']['sendMail']:
            sendMail (conf['mail']['mailId'],
                      conf['mail']['mailPass'],
                      conf['mail']['toEmail'],
                      mail_msg,
                      purchaseOrderFilename)
            print("sendMail success!!")


if __name__ == "__main__":
    for i in range(1):
        # print('process', str(i + 1))
        main()
